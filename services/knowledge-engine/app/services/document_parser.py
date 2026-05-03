"""Document text extraction — PDF, DOCX, HTML, Markdown, plain text."""

from __future__ import annotations

import io
import logging

logger = logging.getLogger(__name__)

SUPPORTED_EXTENSIONS = {
    ".txt", ".md", ".pdf", ".docx", ".html", ".htm", ".srt",
    ".csv", ".xlsx", ".xls",
}


def detect_content_type(filename: str) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    mapping = {
        "txt": "text",
        "md": "markdown",
        "pdf": "pdf",
        "docx": "docx",
        "html": "html",
        "htm": "html",
        "srt": "srt",
        "csv": "csv",
        "xlsx": "xlsx",
        "xls": "xls",
    }
    return mapping.get(ext, "text")


def extract_text(content: bytes, filename: str) -> str:
    """Extract plain text from file bytes based on filename extension."""
    content_type = detect_content_type(filename)
    extractors = {
        "text": _extract_plain,
        "markdown": _extract_plain,
        "pdf": _extract_pdf,
        "docx": _extract_docx,
        "html": _extract_html,
        "srt": _extract_srt,
        "csv": _extract_csv,
        "xlsx": _extract_xlsx,
        "xls": _extract_xls,
    }
    extractor = extractors.get(content_type, _extract_plain)
    try:
        return extractor(content)
    except Exception:
        logger.exception("Failed to extract text from %s, falling back to plain", filename)
        return _extract_plain(content)


def _extract_plain(content: bytes) -> str:
    for encoding in ("utf-8", "gbk", "latin-1"):
        try:
            return content.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            continue
    return content.decode("utf-8", errors="replace")


def _extract_pdf(content: bytes) -> str:
    import fitz  # PyMuPDF

    doc = fitz.open(stream=content, filetype="pdf")
    pages = []
    for page in doc:
        pages.append(page.get_text("text"))
    doc.close()
    return "\n\n".join(pages)


def _extract_docx(content: bytes) -> str:
    from docx import Document

    doc = Document(io.BytesIO(content))
    return "\n\n".join(para.text for para in doc.paragraphs if para.text.strip())


def _extract_html(content: bytes) -> str:
    try:
        from trafilatura import extract

        text = extract(content.decode("utf-8", errors="replace"))
        if text:
            return text
    except Exception:
        pass
    return _extract_plain(content)


def _extract_srt(content: bytes) -> str:
    import pysrt

    subs = pysrt.from_string(_extract_plain(content))
    return "\n".join(f"[{sub.start} --> {sub.end}] {sub.text}" for sub in subs)


def _extract_csv(content: bytes) -> str:
    """CSV → markdown table（保留表头结构，便于 LLM 解析）."""
    import csv as _csv

    text = _extract_plain(content)
    reader = _csv.reader(text.splitlines())
    rows = [r for r in reader if any(c.strip() for c in r)]
    if not rows:
        return ""
    return _rows_to_markdown(rows)


def _extract_xlsx(content: bytes) -> str:
    """XLSX → 每个 sheet 转 markdown 表格，sheet 名作为标题."""
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    parts: list[str] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v) for v in row]
            if any(c.strip() for c in cells):
                rows.append(cells)
        if not rows:
            continue
        parts.append(f"## Sheet: {sheet}\n\n{_rows_to_markdown(rows)}")
    wb.close()
    return "\n\n".join(parts)


def _extract_xls(content: bytes) -> str:
    """老 .xls 二进制 → markdown 表格（xlrd 1.2 仅支持 .xls）."""
    import xlrd

    wb = xlrd.open_workbook(file_contents=content)
    parts: list[str] = []
    for sheet in wb.sheets():
        rows: list[list[str]] = []
        for r in range(sheet.nrows):
            cells = [str(sheet.cell_value(r, c)) for c in range(sheet.ncols)]
            if any(c.strip() for c in cells):
                rows.append(cells)
        if not rows:
            continue
        parts.append(f"## Sheet: {sheet.name}\n\n{_rows_to_markdown(rows)}")
    return "\n\n".join(parts)


def _rows_to_markdown(rows: list[list[str]]) -> str:
    """把二维表格渲染成 markdown 表格（首行当表头）.

    LLM 对 markdown 表格的解析比 CSV 文本更稳定（表头-数据对应清晰）.
    """
    if not rows:
        return ""
    width = max(len(r) for r in rows)
    norm = [r + [""] * (width - len(r)) for r in rows]
    header = norm[0]
    sep = ["---"] * width
    body = norm[1:]
    lines = [
        "| " + " | ".join(_md_cell(c) for c in header) + " |",
        "| " + " | ".join(sep) + " |",
    ]
    for row in body:
        lines.append("| " + " | ".join(_md_cell(c) for c in row) + " |")
    return "\n".join(lines)


def _md_cell(value: str) -> str:
    """转义 markdown 表格里的 | 和换行."""
    s = (value or "").replace("|", "\\|").replace("\n", " ").replace("\r", "")
    return s.strip() or " "
