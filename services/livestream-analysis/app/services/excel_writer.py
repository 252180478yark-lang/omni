"""Excel 报告生成器 — 双 Sheet 结构化报告"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.models import AnalysisResult, Segment, Summary

PHASE_COLORS = {
    "开场暖场": "FFF2CC", "产品介绍": "D9E2F3", "功能演示": "E2EFDA",
    "促单逼单": "FCE4D6", "互动答疑": "E8D5E8", "福利发放": "FFD9D9",
    "过渡衔接": "F2F2F2", "收尾": "D6E4F0",
}
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=10)
DATA_FONT = Font(name="微软雅黑", size=9)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=14, color="2F5496")
SECTION_FONT = Font(name="微软雅黑", bold=True, size=11, color="2F5496")
THIN_BORDER = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="top", wrap_text=True)


class ExcelWriter:

    HEADERS = [
        "时间段", "流程阶段", "画面描述", "背景元素", "贴片元素",
        "人数", "角色",
        "话术-人物A", "话术-人物B", "话术-人物C",
        "语速", "节奏特征", "风格标签", "时长(秒)", "备注",
    ]
    COL_W = [14, 12, 30, 32, 32, 6, 12, 40, 40, 30, 8, 28, 18, 8, 28]

    def write(self, result: AnalysisResult, output_path: str) -> str:
        wb = Workbook()
        self._timeline(wb, result.segments)
        self._summary(wb, result.summary)
        wb.save(output_path)
        return output_path

    def _timeline(self, wb: Workbook, segs: list[Segment]) -> None:
        ws = wb.active
        ws.title = "时间轴分析"

        for ci, h in enumerate(self.HEADERS, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill, c.font, c.alignment, c.border = HEADER_FILL, HEADER_FONT, CENTER, THIN_BORDER

        for ri, seg in enumerate(segs, 2):
            bg_text = "\n".join(f"• {e}" for e in seg.background_elements) if seg.background_elements else ""
            ov_text = "\n".join(f"• {e}" for e in seg.overlay_elements) if seg.overlay_elements else ""
            vals = [
                f"{seg.time_start} ~ {seg.time_end}", seg.phase, seg.visual_description,
                bg_text, ov_text,
                seg.person_count, " / ".join(seg.person_roles),
                self._script(seg, "person_a"), self._script(seg, "person_b"), self._script(seg, "person_c"),
                seg.speech_pace, seg.rhythm_notes, " / ".join(seg.style_tags),
                seg.duration_seconds, seg.notes or "",
            ]
            center_cols = {6, 11, 14}  # 人数、语速、时长(秒)
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.font, c.border = DATA_FONT, THIN_BORDER
                c.alignment = CENTER_WRAP if ci == 1 else (CENTER if ci in center_cols else WRAP)
            color = PHASE_COLORS.get(seg.phase, "FFFFFF")
            ws.cell(row=ri, column=2).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

        for ci, w in enumerate(self.COL_W, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = "A2"
        if segs:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(self.HEADERS))}{len(segs) + 1}"
        ws.row_dimensions[1].height = 28
        for ri in range(2, len(segs) + 2):
            ws.row_dimensions[ri].height = 80

    def _summary(self, wb: Workbook, s: Summary) -> None:
        ws = wb.create_sheet("总结概览")
        row = 1
        ws.merge_cells("A1:D1")
        c = ws.cell(row=1, column=1, value="直播切片分析报告")
        c.font = TITLE_FONT
        c.alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 40
        row = 3

        row = self._sec(ws, row, "基础信息")
        for label, val in [("总时长", s.total_duration), ("分析段数", str(s.total_segments)), ("整体风格", s.overall_style)]:
            ws.cell(row=row, column=1, value=label).font = Font(name="微软雅黑", bold=True, size=10)
            ws.cell(row=row, column=2, value=val).font = DATA_FONT
            ws.cell(row=row, column=1).border = ws.cell(row=row, column=2).border = THIN_BORDER
            row += 1
        row += 1

        if s.person_summary:
            row = self._sec(ws, row, "出场人物")
            for ps in s.person_summary:
                ws.cell(row=row, column=1, value=ps.role).font = Font(name="微软雅黑", bold=True, size=10)
                ws.cell(row=row, column=2, value=ps.description).font = DATA_FONT
                ws.cell(row=row, column=2).alignment = WRAP
                ws.cell(row=row, column=1).border = ws.cell(row=row, column=2).border = THIN_BORDER
                row += 1
            row += 1

        if s.phase_distribution:
            row = self._sec(ws, row, "流程阶段时间分布")
            for ci, lbl in enumerate(["阶段", "时长(秒)", "占比"], 1):
                c = ws.cell(row=row, column=ci, value=lbl)
                c.fill, c.font, c.alignment, c.border = HEADER_FILL, HEADER_FONT, CENTER, THIN_BORDER
            row += 1
            total = sum(s.phase_distribution.values()) or 1
            for phase, sec in s.phase_distribution.items():
                ws.cell(row=row, column=1, value=phase).font = DATA_FONT
                ws.cell(row=row, column=2, value=sec).font = DATA_FONT
                ws.cell(row=row, column=3, value=f"{sec/total*100:.1f}%").font = DATA_FONT
                color = PHASE_COLORS.get(phase, "FFFFFF")
                ws.cell(row=row, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                for ci in range(1, 4):
                    ws.cell(row=row, column=ci).border = THIN_BORDER
                    ws.cell(row=row, column=ci).alignment = CENTER
                row += 1
            row += 1

        for title, items in [("话术亮点 TOP 5", s.highlights), ("改进建议 TOP 5", s.improvements)]:
            if items:
                row = self._sec(ws, row, title)
                for i, item in enumerate(items[:5], 1):
                    ws.cell(row=row, column=1, value=f"{i}.").font = Font(name="微软雅黑", bold=True, size=10)
                    ws.cell(row=row, column=2, value=item).font = DATA_FONT
                    ws.cell(row=row, column=2).alignment = WRAP
                    row += 1
                row += 1

        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 12
        ws.column_dimensions["D"].width = 12

    def _sec(self, ws, row: int, title: str) -> int:
        ws.merge_cells(f"A{row}:D{row}")
        ws.cell(row=row, column=1, value=f"■ {title}").font = SECTION_FONT
        ws.row_dimensions[row].height = 26
        return row + 1

    @staticmethod
    def _script(seg: Segment, key: str) -> str:
        sc = seg.scripts.get(key)
        if sc and sc.content:
            return f"[{sc.role}] {sc.content}" if sc.role else sc.content
        return ""
