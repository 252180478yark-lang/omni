"""千川 CSV/Excel 解析：UTF-8/GBK 自适应，中文列名映射，百分比归一化。"""

from __future__ import annotations

import csv
import io
import re as _re
from typing import Any

import chardet
from openpyxl import load_workbook

# 千川导出列名 → 系统字段的完整映射表
# 按你的实际导出文件逐列匹配
COLUMN_MAPPING: dict[str, str] = {
    # ── 素材标识 ──
    "素材ID": "name",
    "素材id": "name",
    "素材名称": "name",
    "素材名": "name",
    # ── 消耗 ──
    "消耗": "cost",
    "消耗(元)": "cost",
    "花费": "cost",
    "花费(元)": "cost",
    # ── 展示 ──
    "展示次数": "impressions",
    "展示数": "impressions",
    # ── 点击率 ──
    "点击率": "ctr",
    "点击率(%)": "ctr",
    # ── 点击次数 ──
    "点击次数": "clicks",
    "点击数": "clicks",
    # ── CPM / 千展费 ──
    "千次展示费用": "cpm",
    "千次展示费用(元)": "cpm",
    "平均千次展示费用": "cpm",
    "平均千次展示费用(元)": "cpm",
    "CPM": "cpm",
    # ── CPC / 点击均价 ──
    "点击均价": "cpc",
    "点击均价(元)": "cpc",
    "平均点击单价": "cpc",
    "CPC": "cpc",
    # ── 转化数 ──
    "转化数": "conversions",
    "直接转化数": "conversions",
    # ── 转化率 ──
    "点击转化率": "conversion_rate",
    "点击转化率(%)": "conversion_rate",
    "转化率": "conversion_rate",
    "转化率(%)": "conversion_rate",
    "总转化率": "conversion_rate",
    "总转化率(%)": "conversion_rate",
    # ── 转化成本 ──
    "转化成本": "cost_per_result",
    "转化成本(元)": "cost_per_result",
    "消耗成本": "cost_per_result",
    # ── 播放次数 ──
    "播放次数": "plays",
    "播放数": "plays",
    "播放量": "plays",
    # ── 成交金额 ──
    "直接支付成交金额": "direct_pay_amount",
    "直接支付成交金额(元)": "direct_pay_amount",
    "成交金额": "direct_pay_amount",
    "成交金额(元)": "direct_pay_amount",
    # ── 直接支付ROI ──
    "直接支付ROI": "direct_pay_roi",
    "ROI": "direct_pay_roi",
    # ── 3秒播放率 ──
    "3秒播放率": "play_3s_rate",
    "3秒播放率(%)": "play_3s_rate",
    # ── 3秒播放数 ──
    "3秒播放数": "play_3s",
    "3秒播放次数": "play_3s",
    # ── 进度播放率（百分比） ──
    "50%进度播放率": "play_50pct_rate",
    "50%进度播放率(%)": "play_50pct_rate",
    "25%进度播放率": "play_25pct_rate",
    "25%进度播放率(%)": "play_25pct_rate",
    "75%进度播放率": "play_75pct_rate",
    "75%进度播放率(%)": "play_75pct_rate",
    # ── 进度播放数 ──
    "25%进度播放数": "play_25pct",
    "50%进度播放数": "play_50pct",
    "75%进度播放数": "play_75pct",
    # ── 完播率 ──
    "完播率": "completion_rate",
    "完播率(%)": "completion_rate",
    "播放完成率": "completion_rate",
    "播放完成率(%)": "completion_rate",
    "完播数": "play_complete",
    # ── 评论 ──
    "评论次数": "comments",
    "评论数": "comments",
    # ── 分享 ──
    "分享次数": "shares_7d",
    "分享数": "shares_7d",
    "7日分享次数": "shares_7d",
    # ── 关注 ──
    "新增关注数": "new_followers",
    # ── A3 ──
    "新增A3": "new_a3",
    "新增a3": "new_a3",
    "7日新增A3": "new_a3",
    "7日新增a3": "new_a3",
    # ── A3成本 ──
    "7日A3成本": "a3_cost",
    "7日A3成本(元)": "a3_cost",
    "7日a3成本": "a3_cost",
    "7日a3成本(元)": "a3_cost",
    "A3成本": "a3_cost",
    # ── A3占比 ──
    "新增A3占比": "a3_ratio",
    "新增a3占比": "a3_ratio",
    # ── 前展 ──
    "前展": "front_impressions",
    # ── 你的 CSV 新增列 ──
    "素材标题": "_skip",
    "素材创建时间": "_skip",
    "10秒播放率": "play_10s_rate",
    "10秒播放率(%)": "play_10s_rate",
    "5s播放率": "play_5s_rate",
    "5s播放率(%)": "play_5s_rate",
    "5秒播放率": "play_5s_rate",
    "5秒播放率(%)": "play_5s_rate",
    "2秒播放率": "play_2s_rate",
    "2秒播放率(%)": "play_2s_rate",
    "有效播放数": "effective_plays",
    "有效播放次数": "effective_plays",
    "点赞次数": "likes",
    "点赞数": "likes",
    "直接成交订单数": "direct_orders",
    "成交订单数": "direct_orders",
    "关联计划数": "_skip",
    "平均千次展现费用": "cpm",
    "平均千次展现费用(元)": "cpm",
    # ── 跳过不报错的列 ──
    "素材投放量": "_skip",
    "素材计划": "_skip",
    "素材计划数": "_skip",
    "素材标签": "_skip",
    "素材时长": "_skip",
    "素材时长(秒)": "_skip",
    "首帧画面": "_skip",
    "素材来源": "_skip",
    "观看评论率": "_skip",
    "直播间观看人数": "_skip",
    "直播商品点击次数": "_skip",
    "直播间新增粉丝数": "_skip",
}

# ── 列名归一化匹配 ──

_NORM_CACHE: dict[str, str] = {}


def _normalize_col(name: str) -> str:
    s = name.strip().replace("\u3000", "").replace(" ", "")
    s = _re.sub(r"[（(][^)）]*[)）]", "", s)
    return s


def _build_norm_mapping() -> dict[str, str]:
    if _NORM_CACHE:
        return _NORM_CACHE
    for cn_name, sys_field in COLUMN_MAPPING.items():
        norm = _normalize_col(cn_name)
        if norm not in _NORM_CACHE:
            _NORM_CACHE[norm] = sys_field
    return _NORM_CACHE


def _match_columns(csv_columns: list[str]) -> dict[str, str]:
    norm_map = _build_norm_mapping()
    actual_mapping: dict[str, str] = {}
    for csv_col in csv_columns:
        cleaned = (csv_col or "").strip()
        if cleaned in COLUMN_MAPPING:
            actual_mapping[cleaned] = COLUMN_MAPPING[cleaned]
            continue
        norm = _normalize_col(cleaned)
        if norm in norm_map:
            actual_mapping[cleaned] = norm_map[norm]

    if "name" not in actual_mapping.values():
        raise ValueError("未识别到素材名称列（需包含：素材名称/素材名/素材ID 等）")
    return actual_mapping


# ── 编码检测 ──

def detect_encoding(raw_bytes: bytes) -> str:
    if raw_bytes.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    result = chardet.detect(raw_bytes)
    encoding = result.get("encoding") or "utf-8"
    if encoding.upper() in ("GB2312", "GB18030", "GBK"):
        return "gbk"
    return encoding


# ── CSV 解析 ──

def parse_csv(raw_bytes: bytes) -> tuple[list[dict[str, Any]], dict[str, str]]:
    if not raw_bytes or not raw_bytes.strip():
        raise ValueError("文件为空")

    encoding = detect_encoding(raw_bytes)
    text = raw_bytes.decode(encoding)
    if text.startswith("\ufeff"):
        text = text.lstrip("\ufeff")

    reader = csv.DictReader(io.StringIO(text))
    csv_columns = reader.fieldnames or []
    if not csv_columns:
        raise ValueError("CSV 无表头")

    actual_mapping = _match_columns(csv_columns)

    rows: list[dict[str, Any]] = []
    for row in reader:
        parsed: dict[str, Any] = {}
        for csv_col, sys_field in actual_mapping.items():
            raw_value = (row.get(csv_col) or "").strip()
            parsed[sys_field] = _parse_value(sys_field, raw_value)
        name_val = parsed.get("name")
        if not name_val:
            continue
        rows.append(parsed)

    if not rows:
        raise ValueError("CSV 无有效数据行")

    return rows, actual_mapping


# ── 统一入口（CSV / Excel） ──

def parse_tabular_file(raw_bytes: bytes, filename: str) -> tuple[list[dict[str, Any]], dict[str, str]]:
    ext = (filename or "").lower().rsplit(".", 1)[-1] if "." in (filename or "") else ""
    if ext == "csv":
        return parse_csv(raw_bytes)
    if ext in ("xlsx", "xlsm"):
        return parse_excel(raw_bytes)
    if ext == "xls":
        raise ValueError("暂不支持 .xls，请先另存为 .xlsx 或 .csv 再导入")
    raise ValueError("仅支持 CSV / XLSX 文件")


# ── Excel 解析 ──

def parse_excel(raw_bytes: bytes) -> tuple[list[dict[str, Any]], dict[str, str]]:
    if not raw_bytes or not raw_bytes.strip():
        raise ValueError("文件为空")

    try:
        wb = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError("Excel 文件无法读取，请检查文件是否损坏") from e

    ws = wb.active
    iterator = ws.iter_rows(values_only=True)
    try:
        header_row = next(iterator)
    except StopIteration as e:
        raise ValueError("Excel 无数据") from e

    csv_columns = [str(c).strip() if c is not None else "" for c in header_row]
    if not csv_columns:
        raise ValueError("Excel 无表头")

    actual_mapping = _match_columns(csv_columns)

    rows: list[dict[str, Any]] = []
    for values in iterator:
        parsed: dict[str, Any] = {}
        row_map = {
            csv_columns[i]: (values[i] if i < len(values) else None)
            for i in range(len(csv_columns))
        }
        for csv_col, sys_field in actual_mapping.items():
            parsed[sys_field] = _parse_value(sys_field, row_map.get(csv_col))
        name_val = parsed.get("name")
        if not name_val:
            continue
        rows.append(parsed)

    if not rows:
        raise ValueError("Excel 无有效数据行")

    return rows, actual_mapping


# ── 值解析 ──

def _parse_value(field: str, raw: Any) -> int | float | str | None:
    if field == "_skip":
        return None
    if raw is None or raw == "" or raw == "--" or raw == "—":
        return None

    if isinstance(raw, (int, float)):
        num = float(raw)
    else:
        raw_s = str(raw).strip()
        if raw_s.endswith("%"):
            raw_s = raw_s[:-1].strip()
        raw_s = raw_s.replace(",", "")
        if not raw_s:
            return None
        try:
            num = float(raw_s)
        except ValueError:
            return raw_s if field == "name" else None

    int_fields = {
        "impressions", "clicks", "front_impressions",
        "shares_7d", "comments", "plays",
        "play_3s", "play_25pct", "play_50pct", "play_75pct",
        "new_a3", "conversions", "play_complete", "new_followers",
        "effective_plays", "likes", "direct_orders",
    }
    decimal_fields = {
        "cost", "cost_per_result", "direct_pay_amount",
        "cpm", "cpc", "a3_cost", "direct_pay_roi",
    }
    pct_fields = {
        "ctr", "completion_rate", "conversion_rate",
        "a3_ratio", "play_3s_rate",
        "play_25pct_rate", "play_50pct_rate", "play_75pct_rate",
        "play_10s_rate", "play_5s_rate", "play_2s_rate",
    }

    try:
        if field in int_fields:
            return int(num)
        if field in decimal_fields:
            return round(float(num), 6)
        if field in pct_fields:
            v = float(num)
            if v > 1:
                v = v / 100.0
            return round(v, 6)
    except (ValueError, OverflowError):
        return None
    return str(raw).strip() if field == "name" else None
